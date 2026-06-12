from __future__ import annotations

import os
import sys
import tempfile
import threading
import time
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from services.worker.shared.ai import AIBatchSummaryResult, AIProviderError, AISummaryResult

WORKER_DIR = Path(__file__).resolve().parents[1] / "translation_processing"
if str(WORKER_DIR) not in sys.path:
    sys.path.append(str(WORKER_DIR))

import worker


class TranslationWorkerConcurrencyTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.temp_dir.cleanup)
        self.input_path = Path(self.temp_dir.name) / "input.xlsx"
        self.output_path = Path(self.temp_dir.name) / "output.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "DFY官方"
        worksheet.append(["序号", "平台", "发表内容", "研究内容", "分类"])
        for row_index in range(1, 13):
            worksheet.append([row_index, "X", f"content {row_index}", None, None])
        workbook.save(self.input_path)

    def test_process_workbook_uses_configured_batch_concurrency(self) -> None:
        lock = threading.Lock()
        current = 0
        peak = 0

        def fake_generate_batch(texts: list[str], trace_context: dict[str, object] | None = None) -> AIBatchSummaryResult:
            nonlocal current, peak
            with lock:
                current += 1
                peak = max(peak, current)
            time.sleep(0.05)
            with lock:
                current -= 1
            return AIBatchSummaryResult(
                summaries=[f"摘要 {text}" for text in texts],
                trace={**(trace_context or {}), "status": "succeeded", "batchSize": len(texts)},
            )

        with patch.dict(
            os.environ,
            {"XIAOYU_AI_MAX_CONCURRENCY": "6", "XIAOYU_AI_BATCH_SIZE": "2"},
            clear=False,
        ):
            with patch.object(worker, "generate_chinese_summaries_batch_with_trace", side_effect=fake_generate_batch):
                result = worker.process_workbook(self.input_path, self.output_path, "task-1", "attempt-1")

        self.assertTrue(result["ok"])
        self.assertEqual(peak, 6)
        workbook = load_workbook(self.output_path, read_only=True)
        worksheet = workbook["DFY官方"]
        self.assertEqual(worksheet.cell(2, 4).value, "摘要 content 1")

    def test_process_workbook_uses_batch_results_for_writeback(self) -> None:
        batch_calls: list[list[str]] = []

        def fake_batch(texts: list[str], trace_context: dict[str, object] | None = None) -> AIBatchSummaryResult:
            batch_calls.append(list(texts))
            return AIBatchSummaryResult(
                summaries=[f"批量摘要 {text}" for text in texts],
                trace={**(trace_context or {}), "status": "succeeded", "batchSize": len(texts)},
            )

        with patch.dict(
            os.environ,
            {
                "XIAOYU_AI_BATCH_SIZE": "4",
                "XIAOYU_AI_MAX_CONCURRENCY": "2",
            },
            clear=False,
        ):
            with patch.object(worker, "generate_chinese_summaries_batch_with_trace", side_effect=fake_batch):
                result = worker.process_workbook(self.input_path, self.output_path, "task-1", "attempt-1")

        self.assertTrue(result["ok"])
        self.assertEqual([len(batch) for batch in batch_calls], [4, 4, 4])
        workbook = load_workbook(self.output_path, read_only=True)
        worksheet = workbook["DFY官方"]
        self.assertEqual(worksheet.cell(2, 4).value, "批量摘要 content 1")
        self.assertEqual(worksheet.cell(13, 4).value, "批量摘要 content 12")

    def test_failed_batch_can_fallback_to_row_processing(self) -> None:
        def fake_batch(texts: list[str], trace_context: dict[str, object] | None = None) -> AIBatchSummaryResult:
            if "content 5" in texts:
                raise AIProviderError(
                    "AI 摘要生成失败：模型服务暂时不可用，请稍后重试。",
                    trace={
                        **(trace_context or {}),
                        "status": "failed",
                        "failureCategory": "provider_error",
                        "batchSize": len(texts),
                    },
                )
            return AIBatchSummaryResult(
                summaries=[f"批量摘要 {text}" for text in texts],
                trace={**(trace_context or {}), "status": "succeeded", "batchSize": len(texts)},
            )

        def fake_row(text: str, trace_context: dict[str, object] | None = None) -> AISummaryResult:
            return AISummaryResult(
                summary=f"单行补救 {text}",
                trace={**(trace_context or {}), "status": "succeeded"},
            )

        with patch.dict(
            os.environ,
            {
                "XIAOYU_AI_BATCH_SIZE": "4",
                "XIAOYU_AI_MAX_CONCURRENCY": "2",
                "XIAOYU_AI_BATCH_FALLBACK_ENABLED": "true",
            },
            clear=False,
        ):
            with patch.object(worker, "generate_chinese_summaries_batch_with_trace", side_effect=fake_batch):
                with patch.object(worker, "generate_chinese_summary_with_trace", side_effect=fake_row):
                    result = worker.process_workbook(self.input_path, self.output_path, "task-1", "attempt-1")

        self.assertTrue(result["ok"])
        event_types = [event["type"] for event in result["events"]]
        self.assertIn("ai_batch_fallback_started", event_types)
        self.assertIn("ai_batch_fallback_completed", event_types)
        workbook = load_workbook(self.output_path, read_only=True)
        worksheet = workbook["DFY官方"]
        self.assertEqual(worksheet.cell(6, 4).value, "单行补救 content 5")

    def test_sensitive_rows_use_local_fallback_without_ai_calls(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "DFY官方"
        worksheet.append(["序号", "平台", "发表内容", "研究内容", "分类"])
        worksheet.append([1, "X", "12月22日，上海，一名女子在地铁上高喊“打倒司法，打倒共产党”。", None, None])
        worksheet.append([2, "X", "ordinary foreign-language report", None, None])
        workbook.save(self.input_path)

        def fake_batch(texts: list[str], trace_context: dict[str, object] | None = None) -> AIBatchSummaryResult:
            return AIBatchSummaryResult(
                summaries=[f"批量摘要 {text}" for text in texts],
                trace={**(trace_context or {}), "status": "succeeded", "batchSize": len(texts)},
            )

        with patch.dict(
            os.environ,
            {
                "XIAOYU_AI_BATCH_SIZE": "4",
                "XIAOYU_AI_MAX_CONCURRENCY": "2",
            },
            clear=False,
        ):
            with patch.object(worker, "generate_chinese_summaries_batch_with_trace", side_effect=fake_batch) as mocked_batch:
                result = worker.process_workbook(self.input_path, self.output_path, "task-1", "attempt-1")

        self.assertTrue(result["ok"])
        self.assertEqual(mocked_batch.call_count, 1)
        self.assertEqual(result["summary"]["issueCount"], 1)
        self.assertEqual(result["summary"]["issues"][0]["row"], 2)
        self.assertEqual(result["summary"]["issues"][0]["code"], "sensitive_content_fallback")
        event_types = [event["type"] for event in result["events"]]
        self.assertIn("sensitive_content_downgraded", event_types)

        workbook = load_workbook(self.output_path, read_only=True)
        worksheet = workbook["DFY官方"]
        self.assertEqual(worksheet.cell(2, 4).value, "该条内容涉及敏感公共舆情，建议人工复核")
        self.assertEqual(worksheet.cell(3, 4).value, "批量摘要 ordinary foreign-language report")


if __name__ == "__main__":
    unittest.main()
