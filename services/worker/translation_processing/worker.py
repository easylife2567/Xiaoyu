#!/usr/bin/env python3
"""CLI entry point for translation-processing worker commands."""

from __future__ import annotations

import argparse
import json
import os
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook

from constants import REQUIRED_HEADERS, SUPPORTED_SUFFIXES

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from services.worker.shared.ai import AIConfigurationError, AIProcessingError, generate_chinese_summary_with_trace
from services.worker.shared.ai import generate_chinese_summaries_batch_with_trace
from sensitive_content import build_sensitive_issue, detect_sensitive_content

DEFAULT_MAX_CONCURRENCY = 6
DEFAULT_BATCH_SIZE = 20
PROGRESS_ROOT = PROJECT_ROOT / ".data" / "translation-processing" / "progress"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def progress_path(task_id: str | None, attempt_id: str | None) -> Path | None:
    if not task_id or not attempt_id:
        return None
    return PROGRESS_ROOT / f"{task_id}.{attempt_id}.json"


def write_progress_snapshot(task_id: str | None, attempt_id: str | None, progress: dict | None) -> None:
    target_path = progress_path(task_id, attempt_id)
    if target_path is None or progress is None:
        return

    target_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile("w", encoding="utf-8", dir=target_path.parent, delete=False) as temp_file:
        json.dump(progress, temp_file, ensure_ascii=False, indent=2)
        temp_name = temp_file.name
    Path(temp_name).replace(target_path)


def create_progress_state(*, task_id: str | None, attempt_id: str | None, total_rows: int) -> dict:
    return {
        "taskId": task_id,
        "attemptId": attempt_id,
        "status": "processing",
        "totalRows": total_rows,
        "completedRows": 0,
        "totalBatches": 0,
        "completedBatches": 0,
        "currentSheet": None,
        "currentActivity": "任务已启动，等待首个工作表…",
        "activeBatches": [],
        "events": [],
        "updatedAt": utc_now(),
    }


def append_progress_event(progress: dict | None, event_type: str, detail: dict | None = None, *, created_at: str | None = None) -> None:
    if progress is None:
        return

    progress_events = progress.setdefault("events", [])
    progress_events.append(
        {
            "type": event_type,
            "createdAt": created_at or utc_now(),
            "detail": detail or {},
        }
    )
    progress["events"] = progress_events[-12:]
    progress["updatedAt"] = created_at or utc_now()


def set_current_activity(progress: dict | None, activity: str) -> None:
    if progress is None:
        return
    progress["currentActivity"] = activity
    progress["updatedAt"] = utc_now()


def set_current_sheet(progress: dict | None, sheet_name: str) -> None:
    if progress is None:
        return
    progress["currentSheet"] = sheet_name
    progress["updatedAt"] = utc_now()


def register_batch(progress: dict | None, *, sheet_name: str, batch_index: int, start_row: int, end_row: int, batch_size: int) -> None:
    if progress is None:
        return
    progress["totalBatches"] = progress.get("totalBatches", 0) + 1
    active_batches = [*progress.get("activeBatches", [])]
    active_batches.append(
        {
            "sheet": sheet_name,
            "batchIndex": batch_index,
            "startRow": start_row,
            "endRow": end_row,
            "batchSize": batch_size,
        }
    )
    progress["activeBatches"] = active_batches
    set_current_activity(progress, f"正在处理 {sheet_name} 第 {start_row}-{end_row} 行批次")


def complete_batch(progress: dict | None, *, batch_index: int, completed_rows: int, activity: str | None = None) -> None:
    if progress is None:
        return
    progress["completedBatches"] = progress.get("completedBatches", 0) + 1
    progress["completedRows"] = progress.get("completedRows", 0) + completed_rows
    progress["activeBatches"] = [
        batch for batch in progress.get("activeBatches", []) if batch.get("batchIndex") != batch_index
    ]
    if activity:
        set_current_activity(progress, activity)
    else:
        progress["updatedAt"] = utc_now()


def mark_progress_completed_rows(progress: dict | None, rows_completed: int, activity: str) -> None:
    if progress is None:
        return
    progress["completedRows"] = progress.get("completedRows", 0) + rows_completed
    set_current_activity(progress, activity)


def count_processable_rows(workbook) -> int:
    source_rows = 0
    for worksheet in workbook.worksheets:
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        if not all(header in headers for header in REQUIRED_HEADERS):
            continue
        source_index = headers.index("发表内容") + 1
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row_index, source_index).value
            if isinstance(value, str) and value.strip():
                source_rows += 1
    return source_rows


def load_max_concurrency() -> int:
    raw_value = os.getenv("XIAOYU_AI_MAX_CONCURRENCY", str(DEFAULT_MAX_CONCURRENCY)).strip()
    try:
        max_concurrency = int(raw_value)
    except ValueError as error:
        raise AIConfigurationError("AI 摘要生成失败：并发数配置无效。") from error

    if max_concurrency <= 0:
        raise AIConfigurationError("AI 摘要生成失败：并发数配置必须大于 0。")

    return max_concurrency


def load_batch_size() -> int:
    raw_value = os.getenv("XIAOYU_AI_BATCH_SIZE", str(DEFAULT_BATCH_SIZE)).strip()
    try:
        batch_size = int(raw_value)
    except ValueError as error:
        raise AIConfigurationError("AI 摘要生成失败：批大小配置无效。") from error

    if batch_size <= 0:
        raise AIConfigurationError("AI 摘要生成失败：批大小配置必须大于 0。")

    return batch_size


def load_batch_fallback_enabled() -> bool:
    raw_value = os.getenv("XIAOYU_AI_BATCH_FALLBACK_ENABLED", "true").strip().lower()
    return raw_value not in {"0", "false", "no", "off"}


def build_retry_events(ai_trace: dict[str, object], *, sheet: str, row: int) -> list[dict]:
    events = []
    for history in ai_trace.get("retryHistory", []):
        events.append(
            {
                "type": "ai_call_retry_scheduled",
                "createdAt": history.get("finishedAt") or utc_now(),
                "detail": {
                    "sheet": sheet,
                    "row": row,
                    "retryAttempt": history.get("retryAttempt"),
                    "maxRetries": ai_trace.get("maxRetries"),
                    "nextDelayMs": history.get("nextDelayMs"),
                    "failureCategory": history.get("failureCategory"),
                },
            }
        )
    return events


def validate_workbook(input_path: Path) -> dict:
    if input_path.suffix.lower() not in SUPPORTED_SUFFIXES:
        return {
            "ok": False,
            "code": "invalid_file_type",
            "message": "仅支持 .xlsx 或 .xlsm Excel 工作簿。",
        }

    workbook = load_workbook(input_path, read_only=True, data_only=False)
    processable_sheets = []
    source_rows = 0

    for worksheet in workbook.worksheets:
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        if all(header in headers for header in REQUIRED_HEADERS):
            processable_sheets.append(worksheet.title)
            source_index = headers.index("发表内容") + 1
            for row_index in range(2, worksheet.max_row + 1):
                value = worksheet.cell(row_index, source_index).value
                if isinstance(value, str) and value.strip():
                    source_rows += 1

    if not processable_sheets:
        return {
            "ok": False,
            "code": "invalid_workbook_structure",
            "message": "缺少可处理的工作表：需要同时包含“发表内容”“研究内容”“分类”列。",
        }

    if source_rows == 0:
        return {
            "ok": False,
            "code": "invalid_workbook_content",
            "message": "未发现可处理的正文内容。",
        }

    return {
        "ok": True,
        "processableSheets": processable_sheets,
        "sourceRows": source_rows,
        "message": "文件已通过校验，可以开始处理。",
    }


def classify(text: str, sheet_name: str) -> str:
    lowered = text.lower()
    if "taiwan" in lowered or "台湾" in text:
        return "台湾问题"
    if any(term in lowered for term in ("wage arrears", "unpaid wages")) or any(
        term in text for term in ("欠薪", "讨薪", "拖欠工资")
    ):
        return "欠薪讨薪相关"
    if any(term in lowered for term in ("netizen", "mocked", "satirized", "commented")) or any(
        term in text for term in ("网友", "吐槽", "评论", "嘲讽")
    ):
        return "国外网友言论" if sheet_name != "DFY官方" else "国内网友言论"
    if any(term in lowered for term in ("school", "gaokao", "student", "education")) or "高考" in text:
        return "教育相关"
    if any(term in lowered for term in ("company", "factory", "market", "economy")):
        return "经济相关"
    return "社会问题"


def process_row(
    *,
    task_id: str | None,
    attempt_id: str | None,
    sheet_name: str,
    row_index: int,
    source_text: str,
) -> dict:
    ai_result = generate_chinese_summary_with_trace(
        source_text,
        trace_context={
            "taskId": task_id,
            "attemptId": attempt_id,
            "sheet": sheet_name,
            "row": row_index,
        },
    )
    return {
        "row": row_index,
        "summary": ai_result.summary,
        "classification": classify(source_text, sheet_name),
        "trace": ai_result.trace,
    }


def build_row_scope(row_index: int) -> str:
    return f"第 {row_index} 行"


def build_batch_scope(start_row: int, end_row: int) -> str:
    return f"第 {start_row}-{end_row} 行批次" if start_row != end_row else build_row_scope(start_row)


def chunk_rows(row_jobs: list[tuple[int, str]], batch_size: int) -> list[list[tuple[int, str]]]:
    return [row_jobs[index : index + batch_size] for index in range(0, len(row_jobs), batch_size)]


def process_batch(
    *,
    task_id: str | None,
    attempt_id: str | None,
    sheet_name: str,
    batch_index: int,
    rows: list[tuple[int, str]],
) -> dict:
    row_numbers = [row_index for row_index, _ in rows]
    texts = [text for _, text in rows]
    start_row = row_numbers[0]
    end_row = row_numbers[-1]
    batch_result = generate_chinese_summaries_batch_with_trace(
        texts,
        trace_context={
            "taskId": task_id,
            "attemptId": attempt_id,
            "sheet": sheet_name,
            "row": start_row if len(rows) == 1 else None,
            "batchIndex": batch_index,
            "batchRows": row_numbers,
            "batchStartRow": start_row,
            "batchEndRow": end_row,
            "scope": build_batch_scope(start_row, end_row),
        },
    )
    return {
        "batchIndex": batch_index,
        "rows": [
            {
                "row": row_index,
                "summary": summary,
                "classification": classify(source_text, sheet_name),
            }
            for (row_index, source_text), summary in zip(rows, batch_result.summaries, strict=True)
        ],
        "trace": batch_result.trace,
    }


def build_sensitive_row_result(
    *,
    sheet_name: str,
    row_index: int,
    source_text: str,
    sensitive_match,
) -> dict:
    return {
        "row": row_index,
        "summary": sensitive_match.fallback_summary,
        "classification": classify(source_text, sheet_name),
        "sensitive": True,
        "sensitiveMatch": sensitive_match,
    }


def build_failure_payload(
    *,
    error: AIProcessingError,
    task_id: str | None,
    attempt_id: str | None,
    sheet_name: str,
    row_index: int,
    ai_calls: list[dict],
    events: list[dict],
) -> dict:
    trace = error.trace or {
        "taskId": task_id,
        "attemptId": attempt_id,
        "sheet": sheet_name,
        "row": row_index,
        "status": "failed",
        "failureCode": error.code,
    }
    ai_calls.append(trace)
    events.extend(build_retry_events(trace, sheet=sheet_name, row=row_index))
    events.append(
        {
            "type": "ai_call_failed",
            "createdAt": trace.get("finishedAt"),
            "detail": {
                "sheet": sheet_name,
                "row": row_index,
                "failureCategory": trace.get("failureCategory"),
                "durationMs": trace.get("durationMs"),
            },
        }
    )
    return {
        "ok": False,
        "code": error.code,
        "message": str(error),
        "retriable": error.retriable,
        "failureCategory": trace.get("failureCategory"),
        "aiCalls": ai_calls,
        "events": events,
    }


def process_sheet_rows(
    *,
    worksheet,
    source_column: int,
    summary_column: int,
    classification_column: int,
    task_id: str | None,
    attempt_id: str | None,
    ai_calls: list[dict],
    events: list[dict],
    processed_rows: int,
    generated_summaries: int,
    classified_rows: int,
    issues: list[dict],
    progress: dict | None,
) -> tuple[int, int, int] | dict:
    max_concurrency = min(load_max_concurrency(), max(1, worksheet.max_row - 1))
    batch_size = load_batch_size()
    batch_fallback_enabled = load_batch_fallback_enabled()
    row_jobs = []
    sensitive_row_results: dict[int, dict] = {}
    for row_index in range(2, worksheet.max_row + 1):
        source_text = worksheet.cell(row_index, source_column).value
        if not isinstance(source_text, str) or not source_text.strip():
            continue

        sensitive_match = detect_sensitive_content(source_text)
        if sensitive_match is not None:
            sensitive_row_results[row_index] = build_sensitive_row_result(
                sheet_name=worksheet.title,
                row_index=row_index,
                source_text=source_text,
                sensitive_match=sensitive_match,
            )
            issues.append(build_sensitive_issue(sheet=worksheet.title, row=row_index, match=sensitive_match))
            events.append(
                {
                    "type": "sensitive_content_downgraded",
                    "createdAt": utc_now(),
                    "detail": {
                        "sheet": worksheet.title,
                        "row": row_index,
                        "reason": sensitive_match.reason,
                        "requiresHumanReview": True,
                    },
                }
            )
            append_progress_event(
                progress,
                "sensitive_content_downgraded",
                {
                    "sheet": worksheet.title,
                    "row": row_index,
                    "reason": sensitive_match.reason,
                    "requiresHumanReview": True,
                },
            )
            mark_progress_completed_rows(
                progress,
                1,
                f"{worksheet.title} 第 {row_index} 行已降级为模板摘要",
            )
            write_progress_snapshot(task_id, attempt_id, progress)
            continue

        row_jobs.append((row_index, source_text))

    row_results: dict[int, dict] = dict(sensitive_row_results)
    batches = chunk_rows(row_jobs, batch_size)
    with ThreadPoolExecutor(max_workers=max_concurrency) as executor:
        pending: dict = {}
        batch_iter = iter(enumerate(batches, start=1))

        for _ in range(min(max_concurrency, len(batches))):
            batch_entry = next(batch_iter, None)
            if batch_entry is None:
                break
            batch_index, batch_rows = batch_entry
            start_row = batch_rows[0][0]
            end_row = batch_rows[-1][0]
            if len(batch_rows) > 1:
                events.append(
                    {
                        "type": "ai_batch_started",
                        "createdAt": utc_now(),
                        "detail": {
                            "sheet": worksheet.title,
                            "batchIndex": batch_index,
                            "startRow": start_row,
                            "endRow": end_row,
                            "batchSize": len(batch_rows),
                        },
                    }
                )
            register_batch(
                progress,
                sheet_name=worksheet.title,
                batch_index=batch_index,
                start_row=start_row,
                end_row=end_row,
                batch_size=len(batch_rows),
            )
            append_progress_event(
                progress,
                "ai_batch_started" if len(batch_rows) > 1 else "ai_call_started",
                {
                    "sheet": worksheet.title,
                    "batchIndex": batch_index,
                    "startRow": start_row,
                    "endRow": end_row,
                    "batchSize": len(batch_rows),
                    "row": start_row,
                },
            )
            write_progress_snapshot(task_id, attempt_id, progress)
            pending[
                executor.submit(
                    process_batch,
                    task_id=task_id,
                    attempt_id=attempt_id,
                    sheet_name=worksheet.title,
                    batch_index=batch_index,
                    rows=batch_rows,
                )
            ] = (batch_index, batch_rows)

        while pending:
            future = next(as_completed(tuple(pending)))
            batch_index, batch_rows = pending.pop(future)
            start_row = batch_rows[0][0]
            end_row = batch_rows[-1][0]
            try:
                batch_result = future.result()
            except AIProcessingError as error:
                ai_calls.append(error.trace)
                events.extend(build_retry_events(error.trace, sheet=worksheet.title, row=start_row))
                if len(batch_rows) > 1:
                    events.append(
                        {
                            "type": "ai_batch_failed",
                            "createdAt": error.trace.get("finishedAt") or utc_now(),
                            "detail": {
                                "sheet": worksheet.title,
                                "batchIndex": batch_index,
                                "startRow": start_row,
                                "endRow": end_row,
                                "batchSize": len(batch_rows),
                                "failureCategory": error.trace.get("failureCategory"),
                            },
                        }
                    )
                    append_progress_event(
                        progress,
                        "ai_batch_failed",
                        {
                            "sheet": worksheet.title,
                            "batchIndex": batch_index,
                            "startRow": start_row,
                            "endRow": end_row,
                            "batchSize": len(batch_rows),
                            "failureCategory": error.trace.get("failureCategory"),
                        },
                        created_at=error.trace.get("finishedAt"),
                    )
                    set_current_activity(progress, f"{worksheet.title} 第 {start_row}-{end_row} 行批次失败")
                    write_progress_snapshot(task_id, attempt_id, progress)
                if batch_fallback_enabled and len(batch_rows) > 1:
                    events.append(
                        {
                            "type": "ai_batch_fallback_started",
                            "createdAt": utc_now(),
                            "detail": {
                                "sheet": worksheet.title,
                                "batchIndex": batch_index,
                                "startRow": start_row,
                                "endRow": end_row,
                            },
                        }
                    )
                    append_progress_event(
                        progress,
                        "ai_batch_fallback_started",
                        {
                            "sheet": worksheet.title,
                            "batchIndex": batch_index,
                            "startRow": start_row,
                            "endRow": end_row,
                        },
                    )
                    set_current_activity(progress, f"正在降级处理 {worksheet.title} 第 {start_row}-{end_row} 行批次")
                    write_progress_snapshot(task_id, attempt_id, progress)
                    fallback_failed = None
                    for row_index, source_text in batch_rows:
                        try:
                            row_result = process_row(
                                task_id=task_id,
                                attempt_id=attempt_id,
                                sheet_name=worksheet.title,
                                row_index=row_index,
                                source_text=source_text,
                            )
                        except AIProcessingError as row_error:
                            fallback_failed = (row_index, row_error)
                            break
                        ai_calls.append(row_result["trace"])
                        events.extend(build_retry_events(row_result["trace"], sheet=worksheet.title, row=row_index))
                        events.append(
                            {
                                "type": "ai_call_succeeded",
                                "createdAt": row_result["trace"].get("finishedAt"),
                                "detail": {
                                    "sheet": worksheet.title,
                                    "row": row_index,
                                    "durationMs": row_result["trace"].get("durationMs"),
                                },
                            }
                        )
                        append_progress_event(
                            progress,
                            "ai_call_succeeded",
                            {
                                "sheet": worksheet.title,
                                "row": row_index,
                                "durationMs": row_result["trace"].get("durationMs"),
                            },
                            created_at=row_result["trace"].get("finishedAt"),
                        )
                        mark_progress_completed_rows(progress, 1, f"已完成 {worksheet.title} 第 {row_index} 行降级补救")
                        write_progress_snapshot(task_id, attempt_id, progress)
                        row_results[row_index] = row_result

                    if fallback_failed is None:
                        events.append(
                            {
                                "type": "ai_batch_fallback_completed",
                                "createdAt": utc_now(),
                                "detail": {
                                    "sheet": worksheet.title,
                                    "batchIndex": batch_index,
                                    "startRow": start_row,
                                    "endRow": end_row,
                                },
                            }
                        )
                        complete_batch(
                            progress,
                            batch_index=batch_index,
                            completed_rows=0,
                            activity=f"{worksheet.title} 第 {start_row}-{end_row} 行批次已降级完成",
                        )
                        append_progress_event(
                            progress,
                            "ai_batch_fallback_completed",
                            {
                                "sheet": worksheet.title,
                                "batchIndex": batch_index,
                                "startRow": start_row,
                                "endRow": end_row,
                            },
                        )
                        write_progress_snapshot(task_id, attempt_id, progress)
                    else:
                        for remaining in pending:
                            remaining.cancel()
                        failed_row_index, row_error = fallback_failed
                        return build_failure_payload(
                            error=row_error,
                            task_id=task_id,
                            attempt_id=attempt_id,
                            sheet_name=worksheet.title,
                            row_index=failed_row_index,
                            ai_calls=ai_calls,
                            events=events,
                        )
                else:
                    for remaining in pending:
                        remaining.cancel()
                    return build_failure_payload(
                        error=error,
                        task_id=task_id,
                        attempt_id=attempt_id,
                        sheet_name=worksheet.title,
                        row_index=start_row,
                        ai_calls=ai_calls,
                        events=events,
                    )
            else:
                ai_calls.append(batch_result["trace"])
                events.extend(build_retry_events(batch_result["trace"], sheet=worksheet.title, row=start_row))
                if len(batch_rows) > 1:
                    events.append(
                        {
                            "type": "ai_batch_succeeded",
                            "createdAt": batch_result["trace"].get("finishedAt"),
                            "detail": {
                                "sheet": worksheet.title,
                                "batchIndex": batch_index,
                                "startRow": start_row,
                                "endRow": end_row,
                                "batchSize": len(batch_rows),
                                "durationMs": batch_result["trace"].get("durationMs"),
                            },
                        }
                    )
                    append_progress_event(
                        progress,
                        "ai_batch_succeeded",
                        {
                            "sheet": worksheet.title,
                            "batchIndex": batch_index,
                            "startRow": start_row,
                            "endRow": end_row,
                            "batchSize": len(batch_rows),
                            "durationMs": batch_result["trace"].get("durationMs"),
                        },
                        created_at=batch_result["trace"].get("finishedAt"),
                    )
                    complete_batch(
                        progress,
                        batch_index=batch_index,
                        completed_rows=len(batch_rows),
                        activity=f"已完成 {worksheet.title} 第 {start_row}-{end_row} 行批次",
                    )
                    write_progress_snapshot(task_id, attempt_id, progress)
                else:
                    events.append(
                        {
                            "type": "ai_call_succeeded",
                            "createdAt": batch_result["trace"].get("finishedAt"),
                            "detail": {
                                "sheet": worksheet.title,
                                "row": start_row,
                                "durationMs": batch_result["trace"].get("durationMs"),
                            },
                        }
                    )
                    append_progress_event(
                        progress,
                        "ai_call_succeeded",
                        {
                            "sheet": worksheet.title,
                            "row": start_row,
                            "durationMs": batch_result["trace"].get("durationMs"),
                        },
                        created_at=batch_result["trace"].get("finishedAt"),
                    )
                    complete_batch(
                        progress,
                        batch_index=batch_index,
                        completed_rows=1,
                        activity=f"已完成 {worksheet.title} 第 {start_row} 行",
                    )
                    write_progress_snapshot(task_id, attempt_id, progress)
                for row_result in batch_result["rows"]:
                    row_results[row_result["row"]] = row_result

            next_batch_entry = next(batch_iter, None)
            if next_batch_entry is not None:
                next_batch_index, next_batch_rows = next_batch_entry
                next_start_row = next_batch_rows[0][0]
                next_end_row = next_batch_rows[-1][0]
                if len(next_batch_rows) > 1:
                    events.append(
                        {
                            "type": "ai_batch_started",
                            "createdAt": utc_now(),
                            "detail": {
                                "sheet": worksheet.title,
                                "batchIndex": next_batch_index,
                                "startRow": next_start_row,
                                "endRow": next_end_row,
                                "batchSize": len(next_batch_rows),
                            },
                        }
                    )
                register_batch(
                    progress,
                    sheet_name=worksheet.title,
                    batch_index=next_batch_index,
                    start_row=next_start_row,
                    end_row=next_end_row,
                    batch_size=len(next_batch_rows),
                )
                append_progress_event(
                    progress,
                    "ai_batch_started" if len(next_batch_rows) > 1 else "ai_call_started",
                    {
                        "sheet": worksheet.title,
                        "batchIndex": next_batch_index,
                        "startRow": next_start_row,
                        "endRow": next_end_row,
                        "batchSize": len(next_batch_rows),
                        "row": next_start_row,
                    },
                )
                write_progress_snapshot(task_id, attempt_id, progress)
                pending[
                    executor.submit(
                        process_batch,
                        task_id=task_id,
                        attempt_id=attempt_id,
                        sheet_name=worksheet.title,
                        batch_index=next_batch_index,
                        rows=next_batch_rows,
                    )
                ] = (next_batch_index, next_batch_rows)

    for row_index in sorted(row_results):
        row_result = row_results[row_index]
        worksheet.cell(row_index, summary_column).value = row_result["summary"]
        worksheet.cell(row_index, classification_column).value = row_result["classification"]
        processed_rows += 1
        generated_summaries += 1
        classified_rows += 1
        if row_result.get("sensitive"):
            continue

        if not row_result["summary"] or not row_result["classification"]:
            issues.append(
                {
                    "sheet": worksheet.title,
                    "row": row_index,
                    "message": "未能生成完整摘要或分类。",
                }
            )

    return processed_rows, generated_summaries, classified_rows


def process_workbook(input_path: Path, output_path: Path, task_id: str | None = None, attempt_id: str | None = None) -> dict:
    workbook = load_workbook(input_path)
    progress = create_progress_state(task_id=task_id, attempt_id=attempt_id, total_rows=count_processable_rows(workbook))
    write_progress_snapshot(task_id, attempt_id, progress)
    issues = []
    ai_calls = []
    events = []
    processed_rows = 0
    generated_summaries = 0
    classified_rows = 0

    for worksheet in workbook.worksheets:
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        if not all(header in headers for header in REQUIRED_HEADERS):
            continue

        source_column = headers.index("发表内容") + 1
        summary_column = headers.index("研究内容") + 1
        classification_column = headers.index("分类") + 1
        events.append(
            {
                "type": "sheet_started",
                "createdAt": utc_now(),
                "detail": {"sheet": worksheet.title},
            }
        )
        set_current_sheet(progress, worksheet.title)
        append_progress_event(progress, "sheet_started", {"sheet": worksheet.title})
        set_current_activity(progress, f"开始处理工作表 {worksheet.title}")
        write_progress_snapshot(task_id, attempt_id, progress)
        sheet_result = process_sheet_rows(
            worksheet=worksheet,
            source_column=source_column,
            summary_column=summary_column,
            classification_column=classification_column,
            task_id=task_id,
            attempt_id=attempt_id,
            ai_calls=ai_calls,
            events=events,
            processed_rows=processed_rows,
            generated_summaries=generated_summaries,
            classified_rows=classified_rows,
            issues=issues,
            progress=progress,
        )
        if isinstance(sheet_result, dict):
            return sheet_result
        processed_rows, generated_summaries, classified_rows = sheet_result

    if issues:
        blocking_issues = [issue for issue in issues if issue.get("blocking", True)]
        if not blocking_issues:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
            progress["status"] = "completed"
            progress["completedRows"] = progress.get("totalRows", processed_rows)
            progress["activeBatches"] = []
            set_current_activity(progress, f"任务已完成，共处理 {processed_rows} 条内容")
            write_progress_snapshot(task_id, attempt_id, progress)
            return {
                "ok": True,
                "summary": {
                    "processedRows": processed_rows,
                    "generatedSummaries": generated_summaries,
                    "classifiedRows": classified_rows,
                    "issueCount": len(issues),
                    "issues": issues,
                },
                "aiCalls": ai_calls,
                "events": events,
            }

        return {
            "ok": False,
            "code": "output_validation_failed",
            "message": "结果校验失败：存在未完成的摘要或分类。",
            "issues": blocking_issues,
            "retriable": True,
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    progress["status"] = "completed"
    progress["completedRows"] = progress.get("totalRows", processed_rows)
    progress["activeBatches"] = []
    set_current_activity(progress, f"任务已完成，共处理 {processed_rows} 条内容")
    write_progress_snapshot(task_id, attempt_id, progress)
    return {
        "ok": True,
        "summary": {
            "processedRows": processed_rows,
            "generatedSummaries": generated_summaries,
            "classifiedRows": classified_rows,
            "issueCount": len(issues),
            "issues": issues,
        },
        "aiCalls": ai_calls,
        "events": events,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("command", choices=["validate", "process"])
    parser.add_argument("--input", required=True)
    parser.add_argument("--output")
    parser.add_argument("--task-id")
    parser.add_argument("--attempt-id")
    args = parser.parse_args()

    if args.command == "validate":
        print(json.dumps(validate_workbook(Path(args.input)), ensure_ascii=False))
        return

    if args.command == "process":
        if not args.output:
            raise SystemExit("--output is required for process")
        result = process_workbook(Path(args.input), Path(args.output), args.task_id, args.attempt_id)
        result["taskId"] = args.task_id
        result["attemptId"] = args.attempt_id
        print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()
